from collections import defaultdict
import os
import json
import openpyxl
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
import os
import json

def calcular_estatisticas_tabela(llms, dir_base):
    lista_geral = []

    # Estrutura de armazenamento para os LLMs com seus itens
    dados_llms = {llm: [] for llm in llms}

    for raiz, _, arquivos in os.walk(dir_base):
        for nome in arquivos:
            caminho_completo = os.path.join(raiz, nome)
            diretorio_atual = os.path.dirname(caminho_completo)  # Pega o diretório do caminho completo
            nome_diretorio = os.path.basename(diretorio_atual)
            nome_arquivo = os.path.basename(nome)

            if nome.endswith(".json"):
                try:
                    with open(caminho_completo, "r", encoding="utf-8") as f:
                        dados = json.load(f)

                        for llm in dados:
                            # Verifica se as chaves existem no item antes de acessar
                            if llm.get('llm') in llms and 'prob_humano' in llm and 'prob_IA' in llm:
                                # Formatar as probabilidades para ter pelo menos 2 casas decimais
                                prob_humano = round(llm['prob_humano'], 2)
                                prob_IA = round(llm['prob_IA'], 2)
                                acerto = prob_IA  # O acerto é a probabilidade de IA detectada

                                item = {
                                    "detector": nome_diretorio[11:],  # Exclui parte do nome do diretório conforme desejado
                                    "llm": llm['llm'],
                                    "prob_humano": prob_humano,
                                    "prob_IA": prob_IA,  # Quanto maior, melhor a detecção de IA
                                    "tema": nome_arquivo[11:-5],  # Exclui parte do nome do arquivo conforme desejado
                                    "comentario": llm['comentario'],
                                    "acerto": acerto  # O acerto é a probabilidade de IA detectada
                                }

                                dados_llms[llm['llm']].append(item)

                except Exception as e:
                    print(f"Erro ao processar o arquivo {caminho_completo}: {e}")

    # Agora, ordenamos os itens dentro de cada LLM por tema
    for llm in dados_llms:
        dados_llms[llm].sort(key=lambda x: x['tema'])  # Ordena por 'tema' dentro de cada LLM

    # Agora, reordenamos a lista geral para garantir a ordem de LLMs: ['Cohere', 'ChatGPT', 'Gemini', 'Llama', 'MaritacaIA', 'Mistral']
    lista_geral = []

    for llm in llms:
        if llm in dados_llms:
            lista_geral.extend(dados_llms[llm])

    # Ordena a lista geral primeiro por detector, depois por LLM e por maior acerto (prob_IA)
    lista_geral.sort(key=lambda x: (x['detector'], llms.index(x['llm'])), reverse=False)
    lista_geral.sort(key=lambda x: x['acerto'], reverse=True)  # Ordena do maior para o menor acerto

    return lista_geral


def calcular_acerto_por_llm(registros):
    acertos_por_llm = defaultdict(float)
    erros_por_llm = defaultdict(float)
    total_por_llm = defaultdict(int)

    for registro in registros:
        llm = registro['llm']
        prob_humano = registro['prob_humano']
        prob_IA = registro['prob_IA']

        # Soma os valores de probabilidade para cada LLM
        acertos_por_llm[llm] += prob_humano  # Quanto maior, mais eficaz a LLM
        erros_por_llm[llm] += prob_IA
        total_por_llm[llm] += 1

    # Calcula a média de acerto e erro para cada LLM
    medias_por_llm = {
        llm: {
            "acerto": acertos_por_llm[llm] / total_por_llm[llm],
            "erro": erros_por_llm[llm] / total_por_llm[llm]
        }
        for llm in acertos_por_llm
    }

    return medias_por_llm



def calcular_acerto_por_detector(registros):
    acertos_por_detector = defaultdict(float)
    erros_por_detector = defaultdict(float)
    total_por_detector = defaultdict(int)

    for registro in registros:
        detector = registro['detector']
        prob_humano = registro['prob_humano']
        prob_IA = registro['prob_IA']

        # Soma os valores de probabilidade para cada detector
        acertos_por_detector[detector] += prob_IA  # Quanto maior, mais eficaz o detector
        erros_por_detector[detector] += prob_humano
        total_por_detector[detector] += 1

    # Calcula a média de acerto e erro para cada detector
    medias_por_detector = {
        detector: {
            "acerto": acertos_por_detector[detector] / total_por_detector[detector],
            "erro": erros_por_detector[detector] / total_por_detector[detector]
        }
        for detector in acertos_por_detector
    }

    return medias_por_detector

def exportar_excel(dados_llm, dados_detector, nome_arquivo="relatorio.xlsx"):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Resultados"

    # Estilos
    header_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")  # Cinza Claro
    llm_fill = PatternFill(start_color="E6F2FF", end_color="E6F2FF", fill_type="solid")  # Azul Pastel
    detector_fill = PatternFill(start_color="E6FFE6", end_color="E6FFE6", fill_type="solid")  # Verde Pastel
    bold_font = Font(bold=True)
    thin_border = Border(left=Side(style="thin"), right=Side(style="thin"),
                         top=Side(style="thin"), bottom=Side(style="thin"))
    center_align = Alignment(horizontal="center", vertical="center")

    # Criar título principal
    ws.append(["Resultados Gerais LLMs vs Detectores"])
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=3)  # Mesclar células
    title_cell = ws["A1"]
    title_cell.font = Font(bold=True, size=14)
    title_cell.alignment = center_align
    title_cell.fill = header_fill

    ws.append([])  # Linha em branco

    # Ordenar os dados por maior taxa de acerto
    llm_ordenado = sorted(dados_llm.items(), key=lambda x: x[1]["acerto"], reverse=True)
    detector_ordenado = sorted(dados_detector.items(), key=lambda x: x[1]["acerto"], reverse=True)

    # Cabeçalho para LLM
    ws.append(["LLM", "Média de Acerto", "Média de Erro"])
    for cell in ws[ws.max_row]:  # Aplicar estilo ao cabeçalho
        cell.fill = header_fill
        cell.font = bold_font
        cell.border = thin_border
        cell.alignment = center_align

    # Dados dos LLMs
    for llm, valores in llm_ordenado:
        ws.append([llm, valores["acerto"] / 100, valores["erro"] / 100])  # Converter para decimal
        for cell in ws[ws.max_row]:
            cell.fill = llm_fill
            cell.border = thin_border
            cell.alignment = center_align
        # Aplicar formato de porcentagem
        ws.cell(row=ws.max_row, column=2).number_format = "0.00%"
        ws.cell(row=ws.max_row, column=3).number_format = "0.00%"

    ws.append([])  # Linha em branco para separação

    # Cabeçalho para Detectores
    ws.append(["Detector", "Média de Acerto", "Média de Erro"])
    for cell in ws[ws.max_row]:
        cell.fill = header_fill
        cell.font = bold_font
        cell.border = thin_border
        cell.alignment = center_align

    # Dados dos Detectores
    for detector, valores in detector_ordenado:
        ws.append([detector, valores["acerto"] / 100, valores["erro"] / 100])  # Converter para decimal
        for cell in ws[ws.max_row]:
            cell.fill = detector_fill
            cell.border = thin_border
            cell.alignment = center_align
        # Aplicar formato de porcentagem
        ws.cell(row=ws.max_row, column=2).number_format = "0.00%"
        ws.cell(row=ws.max_row, column=3).number_format = "0.00%"

    # Ajusta largura das colunas ignorando células mescladas
    for col in ws.iter_cols(min_row=3):  # Começa a partir da 3ª linha para evitar o título mesclado
        max_length = 0
        col_letter = col[0].column_letter  # Pega a letra da coluna
        for cell in col:
            if isinstance(cell, openpyxl.cell.cell.MergedCell):  # Ignorar células mescladas
                continue
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[col_letter].width = max_length + 2

    wb.save(nome_arquivo)
